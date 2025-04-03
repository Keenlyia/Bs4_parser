import os
from itertools import product

from openpyxl import load_workbook, Workbook
import requests
from bs4 import BeautifulSoup


URL = 'https://rozetka.com.ua/apple-iphone-15-128gb-black/p395460480/'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'uk-UA,uk;q=0.9,en;q=0.8',
    'Referer': 'https://rozetka.com.ua/',
    'Connection': 'keep-alive',
}

# Fetch data from the page
response = requests.get(URL, headers=headers)
soup = BeautifulSoup(response.text, 'html.parser')


def create_or_load_excel(template_path):
    """ Function to create a new file or load an existing one."""
    if os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Color", "Memory", "Seller", "Price", "Discount Price", "Images", "Product Code", "Reviews", "Characteristics"])  # Add header row
        wb.save(template_path)
    return wb, ws


def save_to_excel(product_info):
    """Function to save product data to an Excel file."""
    template_path = "templates/bs4_template.xlsx"  # Specify the path to your Excel template
    wb, ws = create_or_load_excel(template_path)

    # Add a new row with the product data
    ws.append([
        product_info.get("title"),
        product_info.get("color"),
        product_info.get("memory"),
        product_info.get("seller"),
        product_info.get("price"),
        product_info.get("discount_price"),
        ', '.join(product_info.get("image_urls", [])),  # Join image URLs with commas
        product_info.get("product_code"),
        product_info.get("reviews"),
        ', '.join([f"{key}: {value}" for key, value in product_info.get("characteristics", {}).items()])  # Join characteristics with commas
    ])

    # Save the changes in the file
    wb.save(template_path)


product = {}


# Collect product data
try:
    product['title'] = soup.find('h1', attrs={'class': 'title__font'}).text.strip()
except AttributeError:
    product['title'] = None

try:
    product['color'] = soup.find('span', attrs={'class': 'bold'}).text.strip()
except AttributeError:
    product['color'] = None

try:
    product['memory'] = soup.select('span.bold')[1].text.strip()
except (AttributeError, IndexError):
    product['memory'] = None

try:
    seller_text = soup.find('span', attrs={'_ngcontent-rz-client-c2548945983': True}).find_next_sibling('span').text.strip()
    product['seller'] = seller_text.replace("Продавець:", "").strip() if seller_text else None
except AttributeError:
    product['seller'] = None

try:
    product['price'] = soup.find('p', attrs={'class': 'product-price__big'}).text.strip()
except AttributeError:
    product['price'] = None

try:
    product['discount_price'] = soup.find('p', attrs={'class': 'product-price__small'}).text.strip()
except AttributeError:
    product['discount_price'] = None

try:
    product_text = soup.find('span', attrs={'class': 'ms-auto color-black-60'}).text.strip()
    product['product_code'] = product_text.replace("Код:", "").strip() if product_text else None
except AttributeError:
    product['product_code'] = None

try:
    review_text = soup.find('a', attrs={
        'href': 'https://rozetka.com.ua/ua/apple-iphone-15-128gb-black/p395460480/comments/'}).text.strip()
    parts = review_text.split()
    product['reviews'] = next((part for part in parts if part.isdigit()), "0")
except AttributeError:
    product['reviews'] = "0 відгуків"

# Collect image URLs
try:
    images = soup.find_all('img', attrs={'class': 'thumbnail-button__picture'})
    product['image_urls'] = [img.get('src') for img in images if img.get('src')]
except AttributeError:
    product['image_urls'] = []

# Collect product characteristics
try:
    keys = soup.find_all('dt', attrs={'class': 'label'})
    keys_array = [key.text.strip() for key in keys]
except Exception as e:
    print(f"Error while extracting keys: {e}")
    keys_array = []

try:
    values = soup.find_all('dd', attrs={'class': 'value'})
    values_array = [value.text.strip() for value in values]
except Exception as e:
    print(f"Error while extracting values: {e}")
    values_array = []

# Create a dictionary of characteristics if there are keys and values
if keys_array and values_array:
    product['characteristics'] = dict(zip(keys_array, values_array))
else:
    product['characteristics'] = {}

# Save the product data to Excel
save_to_excel(product)

# Output the result
for key, value in product.items():
    print('=' * 50)
    print(f'{key}: {value}')
